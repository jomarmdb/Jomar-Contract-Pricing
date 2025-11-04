import streamlit as st
import pandas as pd
import pdfplumber
import re
from datetime import date
from io import BytesIO
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image

# -----------------------------------------------------------
# CONFIG
# -----------------------------------------------------------
st.set_page_config(page_title="Jomar Price Sheet Creator", layout="wide")

st.markdown("""
<style>
@font-face {
    font-family: 'Proxima Nova';
    src: url('Proxima Nova Font.ttf') format('truetype');
    font-weight: 400;
    font-style: normal;
}

/* Optional bold/italic faces if you have them later:
@font-face {
    font-family: 'Proxima Nova';
    src: url('Proxima Nova Bold.ttf') format('truetype');
    font-weight: 700;
    font-style: normal;
}
*/

/* Apply Proxima Nova globally */
html, body, [class*="st-"], [data-testid="stAppViewContainer"] * {
    font-family: 'Proxima Nova', 'Segoe UI', system-ui, -apple-system, sans-serif !important;
}

/* Optional: Title and headers a little heavier */
h1, h2, h3 {
    font-family: 'Proxima Nova', 'Segoe UI', system-ui, -apple-system, sans-serif !important;
    font-weight: 600 !important;
}
</style>
""", unsafe_allow_html=True)


st.markdown("""
<style>
h1 {
	font-size: 54px !important;
}
</style>
""", unsafe_allow_html=True)

st.markdown(
	"""
	<style>
	@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600&display=swap');

	html, body, [class*="css"]  {
		font-family: 'Montserrat', sans-serif;
	}

	h1, h2, h3 {
		font-family: 'Montserrat', sans-serif;
		font-weight: 600;
	}
	</style>
	""",
	unsafe_allow_html=True
)

BASE_DIR = Path(__file__).parent
PRODUCTS_PATH = BASE_DIR / "JomarList_10272025.xlsx"  # make sure name matches
FLAT_SHEET_NAME = "Jomar List Pricing"
GROUP_SHEET_NAME = "Model Group"
HEADER_ROW_INDEX = 0  # headers start on row 1 in Excel

CODE_MAP = {
	"P": "PART",
	"U": "SUBLINE",
	"S": "SUBGROUP",
	"L": "LINE",
	"G": None,  # ignore group-level contracts
}

# -----------------------------------------------------------
# HELPERS
# -----------------------------------------------------------

def norm_key(val):
	if pd.isna(val):
		return None
	s = str(val).strip()
	if not s:
		return None
	s = (
		s.replace("–", "-")
		 .replace("—", "-")
		 .replace("-", "-")
		 .replace("’", "'")
	)
	return s.upper()


def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
	df = df.copy()
	df.columns = df.columns.str.strip()
	return df


def normalize_flat(df: pd.DataFrame) -> pd.DataFrame:
	df = normalize_cols(df)
	rename_map = {}

	if "Part #" not in df.columns:
		if "Part#" in df.columns:
			rename_map["Part#"] = "Part #"
		elif "Part Number" in df.columns:
			rename_map["Part Number"] = "Part #"
		elif "Part No" in df.columns:
			rename_map["Part No"] = "Part #"

	if "List Price" not in df.columns:
		for col in df.columns:
			if "List Price" in str(col):
				rename_map[col] = "List Price"
				break
		else:
			if "List" in df.columns:
				rename_map["List"] = "List Price"
			elif "Price" in df.columns:
				rename_map["Price"] = "List Price"

	return df.rename(columns=rename_map)


def normalize_model(df: pd.DataFrame) -> pd.DataFrame:
	df = normalize_cols(df)
	rename_map = {}

	if "Part #" not in df.columns and "Part#" in df.columns:
		rename_map["Part#"] = "Part #"

	if "Sub-Group" not in df.columns:
		if "Sub Group" in df.columns:
			rename_map["Sub Group"] = "Sub-Group"
		elif "Subgroup" in df.columns:
			rename_map["Subgroup"] = "Sub-Group"

	if "Sub-Line" not in df.columns:
		if "Sub Line" in df.columns:
			rename_map["Sub Line"] = "Sub-Line"
		elif "Subline" in df.columns:
			rename_map["Subline"] = "Sub-Line"

	if "Line" not in df.columns and "Line " in df.columns:
		rename_map["Line "] = "Line"

	return df.rename(columns=rename_map)


@st.cache_data
def load_product_workbook(path: Path):
	xls = pd.ExcelFile(path)
	flat = pd.read_excel(xls, sheet_name=FLAT_SHEET_NAME, header=0)
	model = pd.read_excel(xls, sheet_name=GROUP_SHEET_NAME, header=0)
	return flat, model

def extract_contract_from_pdf(pdf_file) -> pd.DataFrame:
    """
    ETNA-tuned parser:
      • If a page has a header, parse BOTH (a) valid rows ABOVE the header and (b) all valid rows BELOW it.
      • If a page has no header (but a header was seen earlier), begin at the first plausible data row and parse downward.
      • Strong validation: requires valid Code (P/U/S/L/G), 2 date-like fields, multiplier 0–1, and product text.
      • Skips page headers/footers and ignores the trailing 'X' checkbox column.
    Returns: [key_value, key_type, start_date, end_date, multiplier, key_norm]
    """
    import re
    rows = []
    header_seen_globally = False

    # === Column bands (x0) tuned to ETNA layout ===
    # Product ~ x≈80–190; Code ~ x≈208; Start ~ x≈252; End ~ x≈302; Multi ~ x≈374 (then 'X' at ~421)
    X_PRODUCT_MAX = 200
    X_CODE_MIN,  X_CODE_MAX  = 200, 245
    X_START_MIN, X_START_MAX = 245, 300
    X_END_MIN,   X_END_MAX   = 300, 360
    X_MULTI_MIN, X_MULTI_MAX = 360, 420   # stop before the trailing "X" (~421)

    def is_date_like(s: str) -> bool:
        if not s:
            return False
        s = s.strip()
        # Accept M/D/YY or M/D/YYYY (and MM/DD variants)
        return re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$", s) is not None

    def normalize_code(raw: str) -> str:
        if not raw:
            return ""
        keep = "".join(ch for ch in raw.upper() if ch.isalpha())
        for ch in keep:
            if ch in ("P", "U", "S", "L", "G"):
                return ch
        return ""

    def parse_multiplier(raw: str):
        if raw is None:
            return None
        s = str(raw).strip().replace(",", ".").replace("O", "0").replace("o", "0")
        if s.startswith("."):
            s = "0" + s
        try:
            val = float(s)
            return val if 0 < val <= 1.0 else None
        except Exception:
            return None

    def extract_row_from_words(ws):
        """Return (product, code, start, end, mult) if this line matches bands/validations; else None."""
        product_parts = [w["text"] for w in ws if w["x0"] < X_PRODUCT_MAX]
        code_parts    = [w["text"] for w in ws if X_CODE_MIN  <= w["x0"] < X_CODE_MAX]
        start_parts   = [w["text"] for w in ws if X_START_MIN <= w["x0"] < X_START_MAX]
        end_parts     = [w["text"] for w in ws if X_END_MIN   <= w["x0"] < X_END_MAX]
        multi_parts   = [w["text"] for w in ws if X_MULTI_MIN <= w["x0"] < X_MULTI_MAX]

        if not product_parts:
            return None

        product = " ".join(product_parts).strip()
        code    = normalize_code(code_parts[0] if code_parts else "")
        start   = start_parts[0] if start_parts else ""
        end     = end_parts[0]   if end_parts   else ""
        mult    = parse_multiplier(multi_parts[0] if multi_parts else None)

        # Keep only strong matches
        if code in ("P", "U", "S", "L", "G") and is_date_like(start) and is_date_like(end) and mult is not None:
            return (product, code, start, end, mult)
        return None

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            words = page.extract_words() or []
            # Group words by line (rounded y 'top')
            line_dict = {}
            for w in words:
                top = round(w["top"])
                line_dict.setdefault(top, []).append(w)

            # --- 1) Find header on this page (if present) and record y cutoff
            page_header_top = None
            for top in sorted(line_dict.keys()):
                ws = sorted(line_dict[top], key=lambda x: x["x0"])
                joined = " ".join(w["text"] for w in ws).lower()
                looks_like_header = (
                    ("product" in joined and "group" in joined and "line" in joined)
                    or ("code" in joined and "start" in joined and "end" in joined and ("multi" in joined or "price" in joined))
                )
                if looks_like_header:
                    page_header_top = top
                    header_seen_globally = True
                    break

            # --- 2) If a header exists on this page: parse *above the header* for carryover rows
            if page_header_top is not None:
                for top in sorted(line_dict.keys()):
                    if top >= page_header_top:
                        break  # stop at header
                    ws = sorted(line_dict[top], key=lambda x: x["x0"])
                    # Skip boilerplate at very top like location lines
                    low_line = " ".join(w["text"] for w in ws).lower()
                    if (
                        low_line.startswith("page ")
                        or "jomar valve" in low_line
                        or "customer price sheet assignment" in low_line
                    ):
                        continue
                    extracted = extract_row_from_words(ws)
                    if extracted:
                        rows.append(extracted)

            # --- 3) If no header here but we've seen one globally, find first plausible data row and start there
            page_cutoff_top = None
            if page_header_top is None and header_seen_globally:
                for top in sorted(line_dict.keys()):
                    ws = sorted(line_dict[top], key=lambda x: x["x0"])
                    extracted = extract_row_from_words(ws)
                    if extracted:
                        page_cutoff_top = top - 1  # begin parsing from here downward
                        break

            # If neither header nor cutoff and no header yet globally, skip this page
            if page_header_top is None and page_cutoff_top is None and not header_seen_globally:
                continue

            # --- 4) Parse rows *below* the header or cutoff
            for top in sorted(line_dict.keys()):
                # enforce starting point
                if page_header_top is not None and top <= page_header_top:
                    continue
                if page_header_top is None and page_cutoff_top is not None and top <= page_cutoff_top:
                    continue

                ws = sorted(line_dict[top], key=lambda x: x["x0"])
                low_line = " ".join(w["text"] for w in ws).lower()

                # Skip obvious headers/footers and boilerplate
                if (
                    low_line.startswith("page ")
                    or "jomar valve" in low_line
                    or "customer price sheet assignment" in low_line
                ):
                    continue

                extracted = extract_row_from_words(ws)
                if extracted:
                    rows.append(extracted)

    # --- Build DataFrame in your app’s expected shape
    if not rows:
        return pd.DataFrame(columns=["key_value", "key_type", "start_date", "end_date", "multiplier", "key_norm"])

    df = pd.DataFrame(rows, columns=["key_value", "code", "start_date", "end_date", "multiplier"])
    df["key_type"] = df["code"].map(CODE_MAP)
    df = df[df["key_type"].notna()]

    df["start_date"] = pd.to_datetime(df["start_date"], errors="coerce")
    df["end_date"]   = pd.to_datetime(df["end_date"], errors="coerce")
    df["multiplier"] = pd.to_numeric(df["multiplier"], errors="coerce")

    df["key_norm"] = df["key_value"].apply(norm_key)
    return df[["key_value", "key_type", "start_date", "end_date", "multiplier", "key_norm"]]

def filter_active(contract_df: pd.DataFrame, as_of: date | None = None) -> pd.DataFrame:
	if as_of is None:
		as_of = date.today()

	def _active(r):
		start_ok = pd.isna(r["start_date"]) or (r["start_date"].date() <= as_of)

		if pd.isna(r["end_date"]):
			end_ok = True
		else:
			end_year = r["end_date"].year
			if end_year < 2000:  # your PDFs use 12/31/1949 as "no end date"
				end_ok = True
			else:
				end_ok = r["end_date"].date() >= as_of

		return start_ok and end_ok

	return contract_df[contract_df.apply(_active, axis=1)]


def apply_contract(
	flat_df: pd.DataFrame,
	contract_df: pd.DataFrame,
	default_mult: float = 0.50,
	list_price_col: str = "List Price",
) -> pd.DataFrame:
	active = filter_active(contract_df).copy()

	# build lookup dicts
	part_map = {}
	subline_map = {}
	subgroup_map = {}
	line_map = {}

	for _, r in active.iterrows():
		key = r.get("key_norm")
		if not key:
			continue
		mult = r["multiplier"]
		if r["key_type"] == "PART":
			part_map[key] = mult
		elif r["key_type"] == "SUBLINE":
			subline_map[key] = mult
		elif r["key_type"] == "SUBGROUP":
			subgroup_map[key] = mult
		elif r["key_type"] == "LINE":
			line_map[key] = mult

	# make sure columns exist
	if "Multiplier" not in flat_df.columns:
		flat_df["Multiplier"] = None
	if "Net Price" not in flat_df.columns:
		flat_df["Net Price"] = None

	flat_df[list_price_col] = pd.to_numeric(flat_df[list_price_col], errors="coerce")

	multipliers = []
	sources = []

	for _, row in flat_df.iterrows():
		part     = norm_key(row.get("Part #"))
		subline  = norm_key(row.get("Sub-Line"))
		subgroup = norm_key(row.get("Sub-Group"))
		line     = norm_key(row.get("Line"))

		# 1) part
		if part and part in part_map:
			m = float(part_map[part])
			multipliers.append(m)
			sources.append(f"PART:{part}")
			continue

		# 2) sub-line
		if subline and subline in subline_map:
			m = float(subline_map[subline])
			multipliers.append(m)
			sources.append(f"SUBLINE:{subline}")
			continue

		# 3) sub-group
		if subgroup and subgroup in subgroup_map:
			m = float(subgroup_map[subgroup])
			multipliers.append(m)
			sources.append(f"SUBGROUP:{subgroup}")
			continue

		# 4) line
		if line and line in line_map:
			m = float(line_map[line])
			multipliers.append(m)
			sources.append(f"LINE:{line}")
			continue

		# 5) default
		multipliers.append(default_mult)
		sources.append(f"DEFAULT:{default_mult:.4f}")

	flat_df["Multiplier"] = multipliers
	flat_df["Net Price"] = flat_df[list_price_col] * flat_df["Multiplier"]
	flat_df["Match Source"] = sources

	return flat_df


from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from openpyxl.drawing.image import Image as XLImage

def to_excel_bytes(
	df_dict: dict[str, pd.DataFrame],
	*,
	add_header_form: bool = True,
	logo_path: Path | None = None,          # e.g., BASE_DIR / "jomar_logo.png"
) -> bytes:
	"""
	Exports sheets with:
	  - optional "customer header" form in rows 2-7 (labels in B, underlines in C:D)
	  - data starts at row 9 (header row = 9)
	  - freeze panes at A10 (top 9 rows)
	  - optional logo placed in frozen area (around M2)
	"""
	output = BytesIO()
	with pd.ExcelWriter(output, engine="openpyxl") as writer:
		for sheet_name, df in df_dict.items():
			# --- write dataframe starting at row 9 (0-based startrow=8) ---
			df.to_excel(
				writer,
				sheet_name=sheet_name,
				index=False,
				startrow=8,   # puts header on row 9
			)

			ws = writer.sheets[sheet_name]

			# Hide gridlines if you like a cleaner look
			ws.sheet_view.showGridLines = False

			# Freeze top 9 rows
			ws.freeze_panes = "A10"

			# ---------- Optional: Customer header form (rows 2-7) ----------
			if add_header_form:
				labels = [
					("B2", "Customer:"),
					("B3", "Attn:"),
					("B4", "Email:"),
					("B5", "Rep:"),
					("B6", "Regional:"),
					("B7", "Regional Email:"),
				]
				# write labels
				for addr, text in labels:
					cell = ws[addr]
					cell.value = text
					cell.font = Font(bold=False, italic=True, color="000000")

				# underline area in C2:D7 via bottom border
				underline = Border(bottom=Side(style="dotted", color="000000"))
				for row in range(2, 8):   # rows 2..7 inclusive
					for col in ("C", "D"):
						ws[f"{col}{row}"].border = underline

			# ---------- Optional: Logo in frozen area (around column M) ----------
			if logo_path and Path(logo_path).exists():
				try:
					img = XLImage(str(logo_path))
					# Set exact dimensions (tweak these to your liking)
					img.width = 324   # width in pixels
					img.height = 112   # height in pixels

					# Anchor placement (you can also try 'L2' or 'N2' to reposition)
					img.anchor = "M2"

					ws.add_image(img)
				except Exception as e:
					print(f"Logo insert failed: {e}")
					pass
			# ---------- Style the header row (row 9) per your current spec ----------
			header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
			header_font = Font(bold=False, italic=True, color="000000")
			header_alignment = Alignment(vertical="center")

			for col_idx, col_name in enumerate(df.columns, start=1):
				# header lives at row 9 after startrow=8
				cell = ws.cell(row=9, column=col_idx)
				cell.fill = header_fill
				cell.font = header_font
				cell.alignment = header_alignment

			# ---------- Column widths (adjust if your order changes) ----------
			col_widths = {
				"A": 16,  # Part #
				"B": 20,  # Model #
				"C": 10,  # Size
				"D": 20,  # Sub-Group
				"E": 20,  # Line
				"F": 20,  # Sub-Line
				"G": 10,  # List Price
				"H": 10,  # Multiplier
				"I": 10,  # Net Price
				"J": 10,  # Case
				"K": 10,  # Carton
				"L": 20,  # Weight
				"M": 14,  # UPC
				"N": 118, # Description
				"O": 26,  # Match Source
			}
			for col_letter, width in col_widths.items():
				ws.column_dimensions[col_letter].width = width

			# ---------- Number formats ----------
			max_row = ws.max_row
			cols_by_name = {name: idx + 1 for idx, name in enumerate(df.columns)}

			# currency for list/net price (these columns start at row 10 for data)
			for price_col in ("List Price", "Net Price"):
				if price_col in cols_by_name:
					col_idx = cols_by_name[price_col]
					for row in range(10, max_row + 1):  # data starts on row 10
						ws.cell(row=row, column=col_idx).number_format = "$#,##0.00"

			# multiplier with 4 decimals
			if "Multiplier" in cols_by_name:
				col_idx = cols_by_name["Multiplier"]
				for row in range(10, max_row + 1):
					ws.cell(row=row, column=col_idx).number_format = "0.0000"

			# UPC as whole numbers
			if "UPC" in cols_by_name:
				col_idx = cols_by_name["UPC"]
				for row in range(10, max_row + 1):
					ws.cell(row=row, column=col_idx).number_format = "0"

	output.seek(0)
	return output.getvalue()
# -----------------------------------------------------------
# UI FLOW
# -----------------------------------------------------------
# --- header with title + logo ---
header_left, header_right = st.columns([3, 1])

with header_left:
	st.title("JOMAR CONTRACT PRICE SHEET CREATOR")

with header_right:
	logo_path = BASE_DIR / "Jomar Valve Logo Red.png"
	if logo_path.exists():
		logo = Image.open(logo_path)
		# show it centered in the right column
		st.image(logo, use_container_width=True)
	else:
		# leave empty space so layout still balances
		st.write("")

# 1) load + prepare workbook (quietly)
# 1) load + prepare workbook (quietly)
# 1) load + prepare workbook (quietly)
try:
	flat_list, model_group = load_product_workbook(PRODUCTS_PATH)
except FileNotFoundError:
	st.error(f"Could not find standardized Excel at `{PRODUCTS_PATH}`.")
	st.stop()

# 2) normalize column names
flat_list = normalize_flat(flat_list)
model_group = normalize_model(model_group)

# 3) find the actual part-number column name in BOTH sheets
def find_part_col(cols):
	for c in cols:
		c_clean = str(c).strip().lower()
		if c_clean in ("part #", "part#", "part number", "part no", "part"):
			return c
	return None

flat_part_col = find_part_col(flat_list.columns)
model_part_col = find_part_col(model_group.columns)

if flat_part_col is None:
	st.error(f"❌ 'Jomar List Pricing' is missing a part column. Columns: {list(flat_list.columns)}")
	st.stop()

if model_part_col is None:
	st.error(f"❌ 'Model Group' is missing a part column. Columns: {list(model_group.columns)}")
	st.stop()

# 4) build normalized join key on BOTH
flat_list["Part_Key"] = flat_list[flat_part_col].apply(norm_key)
model_group["Part_Key"] = model_group[model_part_col].apply(norm_key)

# 5) first merge (what we had before)
flat_merged = flat_list.merge(
	model_group[["Part_Key", "Sub-Group", "Line", "Sub-Line"]],
	on="Part_Key",
	how="left"
)

# 6) 🟡 BACKFILL from model_group if merge left blanks
#    (this is the part that fixes your current issue)
mg_subgroup_map = dict(zip(model_group["Part_Key"], model_group.get("Sub-Group")))
mg_line_map     = dict(zip(model_group["Part_Key"], model_group.get("Line")))
mg_subline_map  = dict(zip(model_group["Part_Key"], model_group.get("Sub-Line")))

# fill Sub-Group
if "Sub-Group" not in flat_merged.columns:
	flat_merged["Sub-Group"] = flat_merged["Part_Key"].map(mg_subgroup_map)
else:
	flat_merged["Sub-Group"] = flat_merged["Sub-Group"].fillna(
		flat_merged["Part_Key"].map(mg_subgroup_map)
	)

# fill Line
if "Line" not in flat_merged.columns:
	flat_merged["Line"] = flat_merged["Part_Key"].map(mg_line_map)
else:
	flat_merged["Line"] = flat_merged["Line"].fillna(
		flat_merged["Part_Key"].map(mg_line_map)
	)

# fill Sub-Line
if "Sub-Line" not in flat_merged.columns:
	flat_merged["Sub-Line"] = flat_merged["Part_Key"].map(mg_subline_map)
else:
	flat_merged["Sub-Line"] = flat_merged["Sub-Line"].fillna(
		flat_merged["Part_Key"].map(mg_subline_map)
	)

# 7) detect list-price column (same as before)
list_price_col = None
for col in flat_merged.columns:
	if "List Price" in str(col):
		list_price_col = col
		break

if list_price_col is None:
	st.error("Could not find a column that contains 'List Price' in the pricing sheet.")
	st.stop()

# --- Custom file uploader styling to match subheader look ---
# --- File uploader section (clean and consistent style) ---
st.subheader("Upload Customer PDF Contract:")

# File uploader directly below the subheader
st.markdown("""
    <style>
    div[data-testid="stFileUploader"] {
        margin-top: -0.75rem !important;
    }
    </style>
""", unsafe_allow_html=True)

pdf_file = st.file_uploader(
    "Upload Customer PDF Contract",
    type=["pdf"],
    key="contract_pdf_uploader",
    label_visibility="collapsed"  # hides the duplicate gray label under subheader
)

if pdf_file is not None:
	# parse PDF
	contract_df = extract_contract_from_pdf(pdf_file)

	st.subheader("Contracted Products:")
	if contract_df.empty:
		st.warning("No contract rows were found under the header. Check the PDF format.")
	else:
		st.dataframe(contract_df, use_container_width=True)

		# 🔧 Default multiplier control (shows above the download button)
		default_mult = st.number_input(
			"BASE MULTIPLIER: Select Base Multiplier and Press Enter Key to Refresh",
			min_value=0.0000,
			max_value=1.0000,
			value=0.5000,       # auto-populates as 0.5000
			step=0.0001,
			format="%.4f",
			help="This value is used when a part has no matching Part/Sub-Line/Sub-Group/Line contract."
		)

		# apply to merged pricing data
		priced_df = apply_contract(
			flat_merged.copy(),
			contract_df,
			default_mult=default_mult,      # 👈 use the UI value
			list_price_col=list_price_col,
		)
			
		# 👇 extra polish: make sure these 3 columns are NOT blank in the final sheet
		priced_df["Sub-Group"] = priced_df["Sub-Group"].fillna(
			priced_df["Part_Key"].map(mg_subgroup_map)
		)
		priced_df["Line"] = priced_df["Line"].fillna(
			priced_df["Part_Key"].map(mg_line_map)
		)
		priced_df["Sub-Line"] = priced_df["Sub-Line"].fillna(
			priced_df["Part_Key"].map(mg_subline_map)
		)

		st.subheader("Sample Price Sheet: First 100 Rows")
		st.dataframe(priced_df.head(100), use_container_width=True)

		# 4) download at bottom
		excel_bytes = to_excel_bytes(
			{"Jomar List Pricing (Priced)": priced_df},
			add_header_form=True,
			logo_path=BASE_DIR / "Jomar Valve Logo Red.png",  # or your “Jomar Valve Logo Red.png”
		)

		st.download_button(
			label="⬇️ Download Excel Net Sheet",
			data=excel_bytes,
			file_name="Jomar List & Net Price Sheet.xlsx",
			mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		)

